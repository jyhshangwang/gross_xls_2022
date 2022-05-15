import os
import loguru
from genericpath import exists
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font
import requests
import time
import datetime
import random

def rand_on():
    delay_lst = [0.1, 0.2, 0.3, 0.4 ,0.5]
    delay = random.choice(delay_lst)
    time.sleep(delay)

def get_stock_datetime():
    reqs_date = requests.get("https://dj.mybank.com.tw/z/zc/zcl/zcl_2330.djhtm")
    soup_date = BeautifulSoup(reqs_date.text, 'html.parser')
    date_tmp = (((soup_date.find('table', {'class': 't01'})).find_all('tr')[7]).find_all('td')[0]).text
    return date_tmp

def get_stock_urls(Stock_Num):

    urls = []
    #urls.append(f'http://kgieworld.moneydj.com/ZXQW/zc/zca/zca_{Stock_Num}.djhtm')
    #urls.append(f'https://kgieworld.moneydj.com/z/zc/zca/zca_{Stock_Num}.djhtm')
    #urls.append(f'http://jsjustweb.jihsun.com.tw/z/zc/zca/zca_{Stock_Num}.djhtm')
    urls.append(f'https://dj.mybank.com.tw/z/zc/zca/zca_{Stock_Num}.djhtm')
    #urls.append(f'https://tw.stock.yahoo.com/quote/{Stock_Num}') # Yahoo
    return urls

@loguru.logger.catch
def get_reqs_data(urls): return [ requests.get(link) for link in urls ]

@loguru.logger.catch
def parse_stock_data(reqs):

    dat_p=0
    dat_v=0
    dat_c=0
    dat_t=0
    #dev=0
    for r in reqs:
        soup = BeautifulSoup(r.text,'html.parser')
        blocks = soup.find_all('table', {'class': 't01'})

        for blk in blocks:
            dat_p = float((((blk.find_all('tr')[1]).find_all('td')[7].text)).replace(',',''))
            #dev   = float((((blk.find_all('tr')[2]).find_all('td')[1].text)).replace(',',''))
            #dat_p = dat_p - dev
            dat_v = float((((blk.find_all('tr')[3]).find_all('td')[7].text)).replace(',',''))
            dat_c = float((((blk.find_all('tr')[13]).find_all('td')[1].text)).replace(',',''))
            dat_t = round((dat_v/dat_c/100),2)
    return dat_p,dat_v,dat_t

@loguru.logger.catch
def parse_stock_data_yahoo(reqs): # for checking the price 

    for r in reqs:
        soup = BeautifulSoup(r.text,'html.parser')
        if   soup.find('span', {'class': 'Fz(32px) Fw(b) Lh(1) Mend(16px) D(f) Ai(c) C($c-trend-up)'}) is not None :
            Yahoo_Price = (soup.find('span', {'class': 'Fz(32px) Fw(b) Lh(1) Mend(16px) D(f) Ai(c) C($c-trend-up)'})).text
        elif soup.find('span', {'class': 'Fz(32px) Fw(b) Lh(1) Mend(16px) D(f) Ai(c) C($c-trend-down)'}) is not None:
            Yahoo_Price = (soup.find('span', {'class': 'Fz(32px) Fw(b) Lh(1) Mend(16px) D(f) Ai(c) C($c-trend-down)'})).text
        elif soup.find('span', {'class': 'Fz(32px) Fw(b) Lh(1) Mend(16px) D(f) Ai(c)'}) is not None:
            Yahoo_Price = (soup.find('span', {'class': 'Fz(32px) Fw(b) Lh(1) Mend(16px) D(f) Ai(c)'})).text
        elif soup.find('span', {'class': 'Fz(32px) Fw(b) Lh(1) Mend(16px)'}) is not None:
            Yahoo_Price = (soup.find('span', {'class': 'Fz(32px) Fw(b) Lh(1) Mend(16px)'})).text
        elif soup.find('span', {'class': 'Fz(32px) Fw(b) Lh(1) Mend(16px) C(#fff) Px(6px) Py(2px) Bdrs(4px) Bgc($c-trend-up)'}) is not None:
            Yahoo_Price = (soup.find('span', {'class': 'Fz(32px) Fw(b) Lh(1) Mend(16px) C(#fff) Px(6px) Py(2px) Bdrs(4px) Bgc($c-trend-up)'})).text
        elif soup.find('span', {'class': 'Fz(32px) Fw(b) Lh(1) Mend(16px) C(#fff) Px(6px) Py(2px) Bdrs(4px) Bgc($c-trend-down)'}) is not None:
            Yahoo_Price = (soup.find('span', {'class': 'Fz(32px) Fw(b) Lh(1) Mend(16px) C(#fff) Px(6px) Py(2px) Bdrs(4px) Bgc($c-trend-down)'})).text
        #print(str(Yahoo_Price))
    return Yahoo_Price


def xls_wb_on(path_xls):
    return openpyxl.load_workbook(path_xls) if os.path.exists(path_xls) else openpyxl.Workbook()

def xls_st_on(obj,flg,st_name,idx):

    st_obj=obj
    st_flag=flg
    for stn in st_obj.sheetnames: st_flag+=1 if stn == st_name else +0
    sheet = st_obj[st_name] if st_flag == 1 else st_obj.create_sheet(st_name,idx)

    return sheet

def cal_init():

    global TYP_VAL,TYP_TRN,date_tmp
    TYP_VAL=2.0
    TYP_TRN=1.0
    date_tmp = get_stock_datetime()

    return 0

@loguru.logger.catch
def cal_avg_price(obj,lst,row):

    st_obj=obj
    day_tmp=lst
    rtmp=row
    sum_tmp=[]
    avg_tmp=[]
    for i in range(len(day_tmp)): sum_tmp.append(0)
    for i in range(len(day_tmp)): avg_tmp.append(0)
    for cnt in range(len(day_tmp)):
        for clm in range(st_obj.max_column,st_obj.max_column-day_tmp[cnt],-1): sum_tmp[cnt]+=(st_obj.cell(row=rtmp, column=clm)).value
        avg_tmp[cnt]=round(sum_tmp[cnt]/day_tmp[cnt],2)

    return avg_tmp

@loguru.logger.catch
def cal_incr_rate(obj,lst,row):

    st_obj=obj
    day_tmp=lst
    rtmp=row
    pri_tmp=[]
    rat_tmp=[]
    for i in range(len(day_tmp)): pri_tmp.append(0)
    for i in range(len(day_tmp)): rat_tmp.append(0)
    today_price = (st_obj.cell(row=rtmp, column=st_obj.max_column)).value
    for cnt in range(len(day_tmp)):
        pri_tmp[cnt]=(st_obj.cell(row=rtmp, column=st_obj.max_column-day_tmp[cnt])).value
        rat_tmp[cnt]=round(float((today_price-pri_tmp[cnt])/pri_tmp[cnt]*100),2) if pri_tmp[cnt] != 0 else 0

    return rat_tmp

@loguru.logger.catch
def cal_slope_rate(obj,num,row):
    
    step=5
    avg_lst=[]
    for cnt in range(num): avg_lst.append((obj.cell(row=row, column=obj.max_column-step*cnt)).value)
    for cnt in range(len(avg_lst)): rat_val = ((avg_lst[cnt]-avg_lst[cnt+1])/avg_lst[cnt+1])*100/step
    print(rat_val)
    return rat_val


# =====          =====
# =====   main   =====
# =====          =====
if __name__ == '__main__':
    loguru.logger.add( f'stock_datalog_{datetime.date.today():%Y%m%d}.log', rotation='1 day', retention='7 days', level='DEBUG')

    TEST_M = 0
    HIDAR_EXCEL = [ 0 , 1 , 1 , 0 ]
    cal_init()

    print('\n===================================')
    print(' Time : '+str(datetime.datetime.today()))
    print('===================================')
    start_time = time.time()

    path_xls = 'C:\\Users\\JS Wang\\Desktop\\test\\tmp.xlsx'
    path_fin = 'C:\\Users\\JS Wang\\Desktop\\test\\gross_all_0115.txt'

    if( TEST_M == 1 ):
        fi_stk = open('C:\\Users\\JS Wang\\Desktop\\test\\gross_all_chk.txt','r')
        fo_stk = open('C:\\Users\\JS Wang\\Desktop\\test\\test_mode_out.txt','w')
        lines = fi_stk.readlines()
        for line in lines:
            STK_NUM = int(line)
            print(str(STK_NUM)+' .. ', end='')
            y_dat = parse_stock_data_yahoo(get_reqs_data(get_stock_urls(str(STK_NUM))))
            print(y_dat)
            print(y_dat, file=fo_stk)
        fi_stk.close()
        fo_stk.close()

    if( HIDAR_EXCEL[0] == 1 ):

        fi_stk = open(path_fin,'r')
        lines = fi_stk.readlines()
    
        STK_CNT = 1
        STK_PRI = []
        STK_VOL = []
        STK_TRR = []

        STK_PRI.append(date_tmp)
        STK_VOL.append(date_tmp)
        STK_TRR.append(date_tmp)

        JS_TMP = [0,0]
        for line in lines:
            STK_NUM = int(line)
            JS_TMP = parse_stock_data(get_reqs_data(get_stock_urls(str(STK_NUM))))
            STK_PRI.append(float(JS_TMP[0]))
            STK_VOL.append(float(JS_TMP[1]))
            STK_TRR.append(float(JS_TMP[2]))

            print(">> No."+str(STK_CNT) + "  ...  "+str(STK_NUM))
            #rand_on()
            STK_CNT+=1

        print('\n\n>> STEP1. Write to  ... '+str(path_xls)+'\n\n')
        wb = xls_wb_on(path_xls)
        print(wb.sheetnames)
        st0 = xls_st_on(wb,0,'Price'   ,0)
        st1 = xls_st_on(wb,0,'Volume'  ,1)
        st2 = xls_st_on(wb,0,'Value'   ,2)
        st3 = xls_st_on(wb,0,'Turnover',3)

        cnt_c0 = st0.max_column
        cnt_c1 = st1.max_column
        cnt_c2 = st2.max_column
        cnt_c3 = st3.max_column

        for r1 in range(1,len(STK_PRI)+1): (st0.cell(row=r1, column=cnt_c0+1)).value = STK_PRI[r1-1]
        for r2 in range(1,len(STK_VOL)+1): (st1.cell(row=r2, column=cnt_c1+1)).value = STK_VOL[r2-1]
        for r3 in range(1,len(STK_PRI)+1): (st2.cell(row=r3, column=cnt_c2+1)).value = STK_PRI[r3-1] if r3 == 1 else round((float(STK_PRI[r3-1])*float(STK_VOL[r3-1])/100000),2)
        for r4 in range(1,len(STK_TRR)+1): (st3.cell(row=r4, column=cnt_c3+1)).value = STK_TRR[r4-1]

        print('\n>> Finished !!\n')
        wb.save(path_xls)

    if( HIDAR_EXCEL[1] == 1 ):

        print('\n\n>> STEP2. Calculate MA from  ... '+str(path_xls)+'\n\n')
        wb = xls_wb_on(path_xls)
        print(wb.sheetnames)
        st0 = xls_st_on(wb,0,'Price',0)
        st4 = xls_st_on(wb,0,'3ma'  ,4)
        st5 = xls_st_on(wb,0,'5ma'  ,5)
        st6 = xls_st_on(wb,0,'10ma' ,6)
        st7 = xls_st_on(wb,0,'20ma' ,7)

        cnt_c0 = st0.max_column
        cnt_c4 = st4.max_column
        cnt_c5 = st5.max_column
        cnt_c6 = st6.max_column
        cnt_c7 = st7.max_column

        for r in range(1,st0.max_row+1):
            day_lst=[ 3, 5,10,20]
            avg_lst=[ 0, 0, 0, 0]
            if r != 1 : avg_lst=cal_avg_price(st0,day_lst,r)
            print(avg_lst)
            (st4.cell(row=r, column=cnt_c4+1)).value = avg_lst[0] if r != 1 else date_tmp
            (st5.cell(row=r, column=cnt_c5+1)).value = avg_lst[1] if r != 1 else date_tmp
            (st6.cell(row=r, column=cnt_c6+1)).value = avg_lst[2] if r != 1 else date_tmp
            (st7.cell(row=r, column=cnt_c7+1)).value = avg_lst[3] if r != 1 else date_tmp

        print('\nFinished !!\n')
        wb.save(path_xls)

    if( HIDAR_EXCEL[2] == 1 ):

        print('\n\n>> STEP3. Calculate Inc. Rate from  ... '+str(path_xls)+'\n\n')
        wb = xls_wb_on(path_xls)
        print(wb.sheetnames)
        st0 = xls_st_on(wb,0,'Price',0)
        stA = xls_st_on(wb,0, 'Inc1',10)
        stB = xls_st_on(wb,0, 'Inc3',11)
        stC = xls_st_on(wb,0, 'Inc5',12)
        stD = xls_st_on(wb,0,'Inc10',13)
        stE = xls_st_on(wb,0,'Inc20',14)
        #stF = xls_st_on(wb,0,'40ma',15)
        #stG = xls_st_on(wb,0,'60ma',16)

        cnt1d = stA.max_column
        cnt2d = stB.max_column
        cnt3d = stC.max_column
        cnt4d = stD.max_column
        cnt5d = stE.max_column
        #cnt6d = stF.max_column
        #cnt7d = stG.max_column

        for r in range(1,st0.max_row+1):
            day_lst=[ 1, 3, 5,10,20]
            rat_lst=[ 0, 0, 0, 0, 0]
            if r!= 1 : rat_lst=cal_incr_rate(st0,day_lst,r)
            print(rat_lst)
            (stA.cell(row=r, column=cnt1d+1)).value = rat_lst[0] if r != 1 else date_tmp
            (stB.cell(row=r, column=cnt2d+1)).value = rat_lst[1] if r != 1 else date_tmp
            (stC.cell(row=r, column=cnt3d+1)).value = rat_lst[2] if r != 1 else date_tmp
            (stD.cell(row=r, column=cnt4d+1)).value = rat_lst[3] if r != 1 else date_tmp
            (stE.cell(row=r, column=cnt5d+1)).value = rat_lst[4] if r != 1 else date_tmp

        print('\nFinished !!\n')
        wb.save(path_xls)

    if( HIDAR_EXCEL[3] == 1 ):

        wb = openpyxl.load_workbook(path_xls)
        st0 = xls_st_on(wb,0,'Price'   ,0)
        st1 = xls_st_on(wb,0,'Volume'  ,1)
        st2 = xls_st_on(wb,0,'Value'   ,2)
        st3 = xls_st_on(wb,0,'Turnover',3)

        #for ra in range(2,st0.max_row+1):
        #    for ca in range(1,st0.max_column+1):
        #        print(str((st0.cell(row=ra, column=ca)).value)+', ', end='')
        CNT_JS=1
        CNT_LP=1
        print('\n 代號    價格   3日線   5日線   2周線    月線    成交值   3日漲幅   周漲幅  2周漲幅   月漲幅      股票                ')
        print('--------------------------------------------------------------------------------------------------------------------------------')
        for js0 in range(2,st2.max_row+1):

            val_lst=[0,0,0,0]
            CNT_VAL=0
            for js1 in range(len(val_lst)):
                val_lst[js1]=(st2.cell(row=js0, column=(st2.max_column-js1))).value
                if( val_lst[js1] > TYP_VAL ): CNT_VAL+=1
            if ( CNT_VAL >= 3 ): # >> 1.

                trn_lst=[0,0,0,0]
                CNT_TRN=0
                for js2 in range(1):
                    trn_lst[js2]=(st3.cell(row=js0, column=st3.max_column)).value
                    if( trn_lst[js2] > TYP_TRN ): CNT_TRN+=1
                if( CNT_TRN >= 1 ): # >> 2.

                    day_lst=[ 3, 5,10,20]
                    avg_lst=[ 0, 0, 0, 0]
                    rat_lst=[ 0, 0, 0, 0]
                    td_price=(st0.cell(row=js0, column=st0.max_column)).value
                    avg_lst=cal_avg_price(st0,day_lst,js0)
                    rat_lst=cal_incr_rate(st0,day_lst,js0)
                    cal_tmp=0
                    for js3 in range(len(day_lst)): cal_tmp+=1 if ((td_price/avg_lst[js3] > 0.9) and (td_price/avg_lst[js3] < 1.8)) else +0 #3. diff ratio

                    if( rat_lst[3] > -10 ): #4. Month increase rate
                        if cal_tmp == 4:
                            #if ( (avg_lst[0] > avg_lst[1]) and (avg_lst[1] > avg_lst[2]) and (avg_lst[2] > avg_lst[3]) and (td_price < 500) ):
                            if ( (avg_lst[0] >= avg_lst[1]) and (td_price < 300) ):
                                if( CNT_LP%15 == 0 ): print(' 代號    價格   3日線   5日線   2周線    月線    成交值   3日漲幅   周漲幅  2周漲幅   月漲幅      股票                ')
                                if( CNT_LP%15 == 0 ): print('--------------------------------------------------------------------------------------------------------------------------------')
                                print('%5s'%str((st0.cell(row=js0, column=2)).value), end='')
                                print('%8s'%str(td_price), end='')
                                print('%8s'%str(round(avg_lst[0],2)), end='')
                                print('%8s'%str(round(avg_lst[1],2)), end='')
                                print('%8s'%str(round(avg_lst[2],2)), end='')
                                print('%8s'%str(round(avg_lst[3],2)), end='')
                                print('%8s'%str((st2.cell(row=js0, column=st2.max_column)).value)+str('億'), end='')
                                print('%9s'%str(rat_lst[0])+'% ', end='')
                                print('%7s'%str(rat_lst[1])+'% ', end='')
                                print('%7s'%str(rat_lst[2])+'% ', end='')
                                print('%7s'%str(rat_lst[3])+'% ', end='')
                                print('{:>5}'.format(str((st0.cell(row=js0, column=3)).value)[:4]), end='')
                                print('{:>9}'.format(str((st0.cell(row=js0, column=4)).value)[:8]), end='')
                                print(' >>> 近3日正在啟動!', end='') if( rat_lst[0]/rat_lst[1] > 0.8 ) else print(end='')
                                print()
                                CNT_LP+=1
            CNT_JS+=1
        wb.save(path_xls)


    end_time = time.time()
    print('\nTake time : '+str(round((end_time-start_time),2))+'(S)')

    print()
    print("       *******             *****        *             *     ***********                         *      ")
    print("     *         *         *       *      * *           *    *                                   *       ")
    print("     *          *       *         *     *  *          *    *                                  *        ")
    print("     *           *     *           *    *   *         *    *                                 *         ")
    print("     *            *    *           *    *    *        *    *                                *          ")
    print("     *            *    *           *    *     *       *    *                               *           ")
    print("     *            *    *           *    *      *      *     ***********      *            *            ")
    print("     *            *    *           *    *       *     *    *                  *          *             ")
    print("     *            *    *           *    *        *    *    *                   *        *              ")
    print("     *           *     *           *    *         *   *    *                    *      *               ")
    print("     *          *       *         *     *          *  *    *                     *    *                ")
    print("     *         *         *       *      *           * *    *                      *  *                 ")
    print("       *******             *****        *             *     ***********            *                   ")
    print()
