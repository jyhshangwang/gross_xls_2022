from bs4 import BeautifulSoup
from openpyxl.workbook.workbook import Workbook
import os
import grequests
import requests
import pyquery
import time
import datetime
from fake_useragent import UserAgent
import random
import openpyxl
import winsound
import gross_class as cls
from alive_progress import alive_bar
import gross_sub as sub
import loguru


def get_date():
    reqs = requests.get('https://dj.mybank.com.tw/z/zc/zcl/zcl_2330.djhtm')
    soup = BeautifulSoup(reqs.text, 'html.parser')
    return (((soup.find('table', {'class': 't01'})).find_all('tr')[7]).find_all('td')[0]).text


# >> path set
path_xls = "C:\\Users\\JS Wang\\Desktop\\2022年07月_每日資料.xlsm"
path_stknum = 'C:\\Users\\JS Wang\\Desktop\\test\\gross_all_0715.txt'
path_fout0 = 'C:\\Users\\JS Wang\\Desktop\\data\\__stock__temp__data_d0.txt'
path_fout1 = 'C:\\Users\\JS Wang\\Desktop\\data\\__stock__temp__data_d1.txt'
path_fout2 = 'C:\\Users\\JS Wang\\Desktop\\data\\__stock__temp__data_d2.txt'
path_fout3 = 'C:\\Users\\JS Wang\\Desktop\\data\\__stock__temp__data_d3.txt'
path_fout4 = 'C:\\Users\\JS Wang\\Desktop\\data\\__stock__temp__data_d4.txt'
date_str = get_date()
date_path = (date_str.replace('111','2022')).replace('/','_')
date_sheet = (date_str[4:9]).replace('/','')
path_final = f'C:\\Users\\JS Wang\\Desktop\\data\\tw_stock_data_{date_path}.txt'



class RevenueInfo:
    def __init__(self,month_lst,revenue_lst,mom_rate_lst,yoy_rate_lst,tyoy_rate_lst):
        self.Month_lst = month_lst
        self.Revenue_lst = revenue_lst
        self.Mom_rate_lst = mom_rate_lst
        self.Yoy_rate_lst = yoy_rate_lst
        self.Tyoy_rate_lst = tyoy_rate_lst
    def __repr__(self):
        return [(self.Month_lst)[0],\
                str((self.Revenue_lst)[0]),\
                str((self.Mom_rate_lst)[0])+'%',\
                str((self.Yoy_rate_lst)[0])+'%',\
                str((self.Tyoy_rate_lst)[0])+'%'\
                ]
    def weighted_revenue(self):
        name = self.Month_lst
        rev = self.Revenue_lst
        ratio = [ 1 for i in range(len(name)) ]
        for i in range(len(name)):
            if name[i][-2:] == '02' : ratio[i] = 10/8
            else                    : ratio[i] = 1
        return [ rev[i]*ratio[i] for i in range(len(rev)) ]
    def get_deviation(self):
        rev = self.weighted_revenue()
        if rev[3]+rev[4]+rev[5] == 0 or rev[4]+rev[5]+rev[6] == 0 :
            return ['-','-']
        else:
            dvo1 = (round((((rev[0]+rev[1]+rev[2])-(rev[3]+rev[4]+rev[5]))/(rev[3]+rev[4]+rev[5])),2))*100
            dvo2 = (round((((rev[1]+rev[2]+rev[3])-(rev[4]+rev[5]+rev[6]))/(rev[4]+rev[5]+rev[6])),2))*100
            dvrms = round((dvo1-dvo2),2)
            return [str(dvo1)+'%',str(dvrms)+'%']
    def get_revenue_cmt(self):
        rev = self.weighted_revenue()
        if  ( rev[0] > rev[1] and rev[1] > rev[2] and rev[2] > rev[3] ): cmt = '營收連3增'
        elif( rev[0] > rev[1] and rev[1] > rev[2] and rev[2] < rev[3] ): cmt = '營收連2增'
        elif( rev[0] > rev[1] and rev[1] < rev[2] and rev[2] < rev[3] ): cmt = '營收月增1'
        elif( rev[0] < rev[1] and rev[1] > rev[2] and rev[2] > rev[3] ): cmt = '營收月-1'
        elif( rev[0] < rev[1] and rev[1] < rev[2] and rev[2] > rev[3] ): cmt = '營收連-2'
        elif( rev[0] < rev[1] and rev[1] < rev[2] and rev[2] < rev[3] ): cmt = '營收連-3'
        else:                                                            cmt = 'NA'
        return cmt
    def get_yoy_cmt(self):
        yoy = self.Yoy_rate_lst
        cnt=0
        for i in range(len(yoy)):
            if i == len(yoy)-1: break
            if yoy[i] < yoy[i+1]: cnt+=1
        if   cnt >  3: yoy_cmt = 'down'
        elif cnt == 0: yoy_cmt = 'up'
        else         : yoy_cmt = '-'
        return yoy_cmt


class DailyInfo:
    def __init__(self, \
                op_price,hi_price,lo_price,td_price, \
                up_dn_price,mx_price,mi_price, \
                pe_ratio,mx_volum,mi_volum,td_volum, \
                dividend, \
                inc_year,inc_5day,inc_1mon,inc_2mon,inc_3mon, \
                roe_rate,stock_cnt,ratio_cmt \
        ):
        self.Op_price = float(op_price.replace(',',''))
        self.Hi_price = float(hi_price.replace(',',''))
        self.Lo_price = float(lo_price.replace(',',''))
        self.Td_price = float(td_price.replace(',',''))
        self.Up_dn_price = float(up_dn_price)
        self.Mx_price = float(mx_price.replace(',',''))
        self.Mi_price = float(mi_price.replace(',',''))
        self.Pe_ratio = float(pe_ratio.replace(',','')) if pe_ratio != 'N/A' else 'NEG'
        self.Mx_volum = int(mx_volum.replace(',',''))
        self.Mi_colum = int(mi_volum.replace(',',''))
        self.Td_volum = int(td_volum.replace(',',''))
        self.Dividend = dividend
        self.Inc_year = inc_year
        self.Inc_5day = inc_5day
        self.Inc_1mon = inc_1mon
        self.Inc_2mon = inc_2mon
        self.Inc_3mon = inc_3mon
        self.Roe_rate = roe_rate
        self.Stock_cnt = float(stock_cnt.replace(',',''))
        self.Ratio_cmt = ratio_cmt.replace('、',' ')
    def __repr__(self):
        return [self.Ratio_cmt,self.Pe_ratio,self.Mx_volum,self.Mi_colum,self.Td_volum,self.Stock_cnt,self.Up_dn_price,\
                self.Mx_price,self.Mi_price,self.Td_price,\
                self.Inc_5day,self.Inc_1mon,self.Inc_2mon,self.Inc_3mon,self.Inc_year\
            ]
    def turnover(self): return round(self.Td_volum/(self.Stock_cnt*100),2)
    def yd_price(self): return self.Td_price-self.Up_dn_price
    def is_jump(self): return self.Op_price-self.yd_price()
    def red_k(self): return self.Td_price-self.Op_price
    def diff_rate(self): return round((self.Hi_price-self.Lo_price)/self.yd_price(),2)
    def inc_1day(self): return str(round(self.Up_dn_price/self.yd_price()*100,2))+'%'
    def val(self): return round(self.Td_price*self.Td_volum/100000,2)


def file_open_init(mode):
    fout0 = open(path_fout0, mode, encoding='UTF-8')
    fout1 = open(path_fout1, mode, encoding='UTF-8')
    fout2 = open(path_fout2, mode, encoding='UTF-8')
    fout3 = open(path_fout3, mode, encoding='UTF-8')
    fout4 = open(path_fout4, mode, encoding='UTF-8')

    return [fout0,fout1,fout2,fout3,fout4]


def file_close(lst):
    for i in range(len(lst)): lst[i].close()


def beep():
    duration = 1680 # mS
    freq = 300      # Hz
    for i in range(1):
        if i%2 == 0 : winsound.Beep(freq, duration)
    time.sleep(0.2)


def get_urls_lst(urls,num):
    if num == 0: urls_lst = [ f'https://dj.mybank.com.tw/z/zc/zch/zch_{url}.djhtm' for url in urls ]
    if num == 1: urls_lst = [ f'http://jsjustweb.jihsun.com.tw/z/zc/zce/zce_{url}.djhtm' for url in urls ]
    if num == 2: urls_lst = [ f'http://jsjustweb.jihsun.com.tw/z/zc/zca/zca_{url}.djhtm' for url in urls ]
    if num == 3: urls_lst = [ f'https://kgieworld.moneydj.com/z/zc/zcl/zcl_{url}.djhtm' for url in urls ]
    if num == 4: urls_lst = [ f'http://jsjustweb.jihsun.com.tw/z/zc/zcn/zcn_{url}.djhtm' for url in urls ]
    return urls_lst


def type_inv(dat): return float((dat.replace('%','')).replace(',','')) if dat != '' else 0

def rmcma(dat): return dat.replace(',','')

def checkString(dat): return '0' if dat == '' else dat

def get_reqs_data_asynch(urls):
    reqs = ( grequests.get(url) for url in urls )
    response = grequests.imap(reqs, grequests.Pool(len(urls)))
    return response


def get_reqs_data(urls):
    reqs = [ requests.get(url) for url in urls ]
    return reqs


@loguru.logger.catch
def parse_data_0(response,cnt_stk): # Month
    total_lst = []
    bar = cls.ProgressBar(cnt_stk)
    for r in response:
        d = pyquery.PyQuery(r.text)
        tits = list(d('title').items())[0].text().strip().replace('個股合併月營收-','')
        tbls = list(d('table').items())
        tbls = tbls[2:3]
        for tbl in tbls:
            trs = list(tbl('tr').items())
            trs = trs[6:13]
            month_lst = []
            revenue_lst = []
            mom_rate_lst = []
            yoy_rate_lst = []
            tyoy_rate_lst = []
            for tr in trs:
                tds = list(tr('td').items())
                month_lst.append(tds[0].text().strip())
                if tds[1].text().strip() == '':
                    revenue_lst.append(0)
                else:
                    revenue_lst.append(int(str(tds[1].text().strip()).replace(',',''))/100000)
                if tds[2].text().strip() == '':
                    mom_rate_lst.append(0)
                else:
                    mom_rate_lst.append(type_inv(checkString(tds[2].text().strip())))
                if tds[4].text().strip() == '':
                    yoy_rate_lst.append(0)
                else:
                    yoy_rate_lst.append(type_inv(checkString(tds[4].text().strip())))
                if tds[6].text().strip() == '':
                    tyoy_rate_lst.append(0)
                else:    
                    tyoy_rate_lst.append(type_inv(checkString(tds[6].text().strip())))
        stk = RevenueInfo(month_lst,revenue_lst,mom_rate_lst,yoy_rate_lst,tyoy_rate_lst)
        tmp_lst = stk.__repr__()
        dev = stk.get_deviation()
        tmp_lst.insert(0,dev[0])
        tmp_lst.insert(1,dev[1])
        tmp_lst.insert(2,stk.get_revenue_cmt())
        tmp_lst.append(stk.get_yoy_cmt())
        str_dat = ';'.join(tmp_lst)
        total_lst.append([int(tits),str_dat])
        bar.update()
    loguru.logger.success('>> Month_info_success.')
    return total_lst


@loguru.logger.catch
def parse_data_1(response,cnt_stk): # Season
    season_lst = []
    bar = cls.ProgressBar(cnt_stk)
    for r in response:
        d = pyquery.PyQuery(r.text)
        tits = (list(d('title').items()))[0].text().strip().replace('個股獲利能力-','')
        tbls = list(d('table').items())
        tbls = tbls[2:3]
        for tbl in tbls:
            trs = list(tbl('tr').items())
            trs = trs[3:7]
            sea_lst = []
            for tr in trs:
                tds = list(tr('td').items())
                sea_lst.append( [ tds[num].text().strip() for num in range(len(tds)) if num == 0 or num == 4 or num == 6 or num == 10 ] )
            for i in range(len(sea_lst[0:1])):
                gross_diff = str(round(type_inv(sea_lst[i][1])-type_inv(sea_lst[i+1][1]), 2))+'%'
                profit_diff = str(round(type_inv(sea_lst[i][2])-type_inv(sea_lst[i+1][2]), 2))+'%'
                eps_rate = round(( float(sea_lst[i][-1]) - float(sea_lst[i+1][-1]) ) / float(sea_lst[i+1][-1]), 2) if float(sea_lst[i+1][-1]) > 0 else 'NEG'
            eps_sum = 0
            for i in range(len(sea_lst)): eps_sum += float(sea_lst[i][-1])
        dat_lst = []
        dat_lst.extend(sea_lst[0])
        dat_lst.insert(2,gross_diff)
        dat_lst.insert(4,profit_diff)
        dat_lst.append(sea_lst[1][-1])
        dat_lst.append(sea_lst[2][-1])
        dat_lst.append(sea_lst[3][-1])
        dat_lst.append(str(eps_rate))
        dat_lst.append(str(eps_sum))
        str_dat = ';'.join(dat_lst)
        season_lst.append([int(tits),str_dat])
        bar.update()
    loguru.logger.success('>> Season_info_success.')
    return season_lst


@loguru.logger.catch
def parse_data_2(response,cnt_stk): # Daily
    daily_lst = []
    bar = cls.ProgressBar(cnt_stk)
    for r in response:
        d = pyquery.PyQuery(r.text)
        tits = (list(d('title').items()))[0].text().strip().replace('個股基本資料-','')
        tbls = list(d('table').items())
        tbls = tbls[2:3]
        for tbl in tbls:
            trs = list(tbl('tr').items())
            trs = trs[1:24]
            trs = trs[0:3]+trs[4:5]+trs[6:11]+trs[12:13]+trs[19:20]
            for tr in trs:
                if tr == trs[0]: # price
                    tds = list(tr('td').items())
                    op_price = tds[1].text().strip()
                    hi_price = tds[3].text().strip()
                    lo_price = tds[5].text().strip()
                    td_price = tds[7].text().strip()
                if tr == trs[1]:
                    tds = list(tr('td').items())
                    up_dn_price = tds[1].text().strip()
                    mx_price = tds[3].text().strip()
                    mi_price = tds[5].text().strip()
                if tr == trs[2]:
                    tds = list(tr('td').items())
                    pe_ratio = tds[1].text().strip()
                    mx_volum = tds[3].text().strip()
                    mi_volum = tds[5].text().strip()
                    td_volum = tds[7].text().strip()
                if tr == trs[3]:
                    tds = list(tr('td').items())
                    dividend = tds[1].text().strip()
                if tr == trs[4]:
                    tds = list(tr('td').items())
                    inc_year = tds[1].text().strip()
                if tr == trs[5]:
                    tds = list(tr('td').items())
                    inc_5day = tds[1].text().strip()
                if tr == trs[6]:
                    tds = list(tr('td').items())
                    inc_1mon = tds[1].text().strip()
                if tr == trs[7]:
                    tds = list(tr('td').items())
                    inc_2mon = tds[1].text().strip()
                if tr == trs[8]:
                    tds = list(tr('td').items())
                    inc_3mon = tds[1].text().strip()
                    roe_rate = tds[5].text().strip()
                if tr == trs[9]:
                    tds = list(tr('td').items())
                    stock_cnt = tds[1].text().strip()
                if tr == trs[10]:
                    tds = list(tr('td').items())
                    ratio_cmt = tds[1].text().strip()
        stk = DailyInfo(op_price,hi_price,lo_price,td_price,up_dn_price,mx_price,mi_price,\
                        pe_ratio,mx_volum,mi_volum,td_volum,dividend,\
                        inc_year,inc_5day,inc_1mon,inc_2mon,inc_3mon,roe_rate,stock_cnt,ratio_cmt)
        dat_lst = []
        dat_lst = stk.__repr__()
        dat_lst.insert(5,'')
        dat_lst.insert(6,'')
        dat_lst.insert(7,stk.turnover())
        dat_lst.insert(12,stk.is_jump())
        dat_lst.insert(13,stk.red_k())
        dat_lst.insert(14,stk.diff_rate())
        dat_lst.insert(16,stk.yd_price())
        dat_lst.insert(17,stk.inc_1day())
        dat_lst.append(stk.val())
        for i in range(len(dat_lst)): dat_lst[i] = str(dat_lst[i])
        str_dat = ';'.join(dat_lst)
        daily_lst.append([int(tits),str_dat])
        bar.update()
    loguru.logger.success('>> Daily_info_success.')
    return daily_lst


@loguru.logger.catch
def parse_data_3(response,stk_cnt): # Counter
    tmp_lst = []
    bar = cls.ProgressBar(stk_cnt)
    for r in response:
        d = pyquery.PyQuery(r.text)
        tits = (list(d('title').items()))[0].text().strip().replace('個股法人持股-','')
        tbls = list(d('table').items())
        tbls = tbls[2:3]
        counter_lst = []
        for tbl in tbls:
            trs = list(tbl('tr').items())
            trs = trs[7:13]
            for tr in trs:
                tds = list(tr('td').items())
                if tds[1].text().strip() != '--':
                    cnt_lst = [ int((tds[i].text().strip()).replace(',','')) for i in range(len(tds)) if 0 < i < 5 ]
                else:
                    cnt_lst = [ 0 for i in range(len(tds)) if 0 < i < 5 ]
                counter_lst.append(cnt_lst)
        cmt = ''
        for i in range(len(counter_lst)-1):
            if   counter_lst[i][1] >  0 : cmt+='+'
            elif counter_lst[i][1] == 0 : cmt+='.'
            else                        : cmt+='-'
        if   cmt[0:5] == '+++++'                 : cmt = '連5買'
        elif cmt[0:4] == '++++' and cmt[4] != '+': cmt = '連4買'
        elif cmt[0:3] == '+++'  and cmt[3] != '+': cmt = '連3買'
        elif cmt[0:2] == '++'   and cmt[2] != '+': cmt = '連2買'
        elif cmt[0:1] == '+'    and cmt[1] != '+': cmt = '連1買'
        elif cmt[0:1] == '-'    and cmt[1] != '-': cmt = '賣1日'
        elif cmt[0:2] == '--'   and cmt[2] != '-': cmt = '賣2日'
        elif cmt[0:3] == '---'  and cmt[3] != '-': cmt = '賣3日'
        elif cmt[0:4] == '----' and cmt[4] != '-': cmt = '賣4日'
        elif cmt[0:5] == '-----'                 : cmt = '賣5日'
        elif cmt[0:5] == '.....'                 : cmt = 'NA'
        else                                     : cmt = 'No rule'

        for i in range(len(counter_lst)):
            for j in range(len(counter_lst[i])):
                counter_lst[i][j] = str(counter_lst[i][j])

        counter_lst = [ ';'.join(counter_lst[i]) for i in range(len(counter_lst)) ]
        str_dat = ';'.join(counter_lst)
        str_dat = str_dat+';'+cmt
        tmp_lst.append([int(tits),str_dat])
        bar.update()
    loguru.logger.success('>> Counter_info_success.')
    return tmp_lst


@loguru.logger.catch
def parse_data_4(response,stk_cnt):
    margin_trad_short_sale_lst = []
    bar = cls.ProgressBar(stk_cnt)
    for r in response:
        d = pyquery.PyQuery(r.text)
        tits = (list(d('title').items())[0]).text().strip().replace('個股融資融券-','')
        tbls = list(d('table').items())
        tbls = tbls[2:3]
        for tbl in tbls:
            trs = list(tbl('tr').items())
            trs = trs[7:13]

            financial_lst = [0,0,0,0,0,0]
            cnt=0
            for tr in trs:
                tds = list(tr('td').items())
                financial_lst[cnt] = [ rmcma(tds[i].text().strip()) for i in range(len(tds)) ]
                cnt+=1
        if financial_lst[0] == 0 or len(financial_lst) != 6:
            margin_trad_short_sale_lst.append([int(tits),'0;0;0;0;0;-'])
            bar.update()
            continue

        def check_rgzratio(lst):
            cnt = 0
            for i in range(len(lst)-1):
                if type_inv(lst[0][13]) < 30: break
                if type_inv(lst[i][13]) > 20: cnt+=1
            cmt = 'High' if cnt == 5 else '-'
            return cmt
        tmp_lst = []
        tmp_lst.append(financial_lst[0][ 5]) # Today margin purchase
        tmp_lst.append(financial_lst[5][ 1]) # Weekly margin purchase
        tmp_lst.append(str(type_inv(financial_lst[0][7])-type_inv(financial_lst[4][7]))+'%') # Weekly margin purchase ratio
        tmp_lst.append(financial_lst[0][12]) # Today short sale
        tmp_lst.append(financial_lst[5][ 3]) # Weekly short sale
        tmp_lst.append(check_rgzratio(financial_lst))
        str_dat = ';'.join(tmp_lst)
        margin_trad_short_sale_lst.append([int(tits),str_dat])
        bar.update()
    print()
    loguru.logger.success('>> Financial_info_success.')
    return margin_trad_short_sale_lst


def main():

    loguru.logger.info(sub.get_stock_datetime())
    start_time = time.time()
    with open(path_stknum,'r') as fin:
        stocks = fin.readlines()
    stock_num_lst = [ int(stock) for stock in stocks ]

    file = file_open_init('w')

    STK_DAT0 = parse_data_0(get_reqs_data_asynch(get_urls_lst(stock_num_lst,0)),len(stock_num_lst))
    for stock in stocks:
        for i in range(len(STK_DAT0)):
            if int(stock) == STK_DAT0[i][0]:
                print(str(STK_DAT0[i][0])+';'+STK_DAT0[i][1], file=file[0])
                break

    STK_DAT1 = parse_data_1(get_reqs_data_asynch(get_urls_lst(stock_num_lst,1)),len(stock_num_lst))
    for stock in stocks:
        for i in range(len(STK_DAT1)):
            if int(stock) == STK_DAT1[i][0]:
                print(str(STK_DAT1[i][0])+';'+STK_DAT1[i][1], file=file[1])
                break

    STK_DAT2 = parse_data_2(get_reqs_data_asynch(get_urls_lst(stock_num_lst,2)),len(stock_num_lst))
    for stock in stocks:
        for i in range(len(STK_DAT2)):
            if int(stock) == STK_DAT2[i][0]:
                print(str(STK_DAT2[i][0])+';'+STK_DAT2[i][1], file=file[2])
                break

    STK_DAT3 = parse_data_3(get_reqs_data_asynch(get_urls_lst(stock_num_lst,3)),len(stock_num_lst))
    for stock in stocks:
        for i in range(len(STK_DAT3)):
            if int(stock) == STK_DAT3[i][0]:
                print(str(STK_DAT3[i][0])+';'+STK_DAT3[i][1], file=file[3])
                break

    STK_DAT4 = parse_data_4(get_reqs_data_asynch(get_urls_lst(stock_num_lst,4)),len(stock_num_lst))
    for stock in stocks:
        for i in range(len(STK_DAT4)):
            if int(stock) == STK_DAT4[i][0]:
                print(str(STK_DAT4[i][0])+';'+STK_DAT4[i][1], file=file[4])
                break

    file_close(file)

    title_str = '股票代號;營收趨勢DVO;營收變化DVRMS;月營收說明;月份;營業額-億;月增率;年增率;總年增率;年增說明;\
                股票代號;季;毛利率;較上季;營益率;較上季;EPS[0]元;EPS[1];EPS[2];EPS[3];Ratio(%);SUM(EPS);\
                股票代號;營收比重;目前本益比;1年內最大量;1年內最小量;今日成交量;昨日成交量;量增率;周轉率;股本(萬張);漲跌;1年最高價;1年最低價;跳空?開;紅K?走;振幅率;\
                今日股價;昨日股價;1日漲幅;5日漲幅;20日漲幅;40日漲幅;60日漲幅;今年以來;今日成交值(億);\
                股票代號;1.外資;1.投信;1.自營;1.總;2.外資;2.投信;2.自營;2.總;3.外資;3.投信;3.自營;3.總;4.外資;4.投信;4.自營;4.總;5.外資;5.投信;5.自營;5.總;近5總.外資;近5總.投信;近5總.自營;近5總.總和;投信動作;\
                股票代號;今融資增減;近5融資增減;變化%值;今融券增減;近5券資比%值;資券說明'


    f_merge = file_open_init('r')
    parse0 = f_merge[0].readlines()
    parse1 = f_merge[1].readlines()
    parse2 = f_merge[2].readlines()
    parse3 = f_merge[3].readlines()
    parse4 = f_merge[4].readlines()

    with open(path_final,'w', encoding='UTF-8') as f:
        print(" ===================", file = f)
        print("   Date:"+date_str, end='', file = f)
        print("\n ===================", file = f)
        print(file = f)
        print(file = f)
        print(title_str, file=f)
        for i in range(len(parse0)):
            total_parse = parse0[i].strip()+';'+\
                          parse1[i].strip()+';'+\
                          parse2[i].strip()+';'+\
                          parse3[i].strip()+';'+\
                          parse4[i].strip()
            print(total_parse, file=f)

    file_close(f_merge)
    loguru.logger.success(f'>> Merge the all txt files successfully. {date_path}')

    try:
        os.remove(path_fout0)
        os.remove(path_fout1)
        os.remove(path_fout2)
        os.remove(path_fout3)
        os.remove(path_fout4)
    except OSError as e:
        loguru.logger.error(e)
    finally:
        loguru.logger.success('>> Remove temp data file successfully.')
    
    with open(path_final, 'r', encoding='UTF-8') as fread:
        SUMSTK = fread.readlines()

    try:
        if os.path.exists(path_xls):
            wb = openpyxl.load_workbook(filename=path_xls, read_only=False, keep_vba=True)
        else:
            wb = openpyxl.Workbook()
        for stn in wb.sheetnames:
            if stn == date_sheet:
                wb.remove(wb[stn])
        wb.create_sheet(title=date_sheet)
        st = wb[date_sheet]
    except OSError as e:
        loguru.logger.error(e)
    finally:
        loguru.logger.success('>> Open the excel file successfully.')

    r_cnt=1
    for d in SUMSTK:
        (st.cell(row=r_cnt, column=1)).value = d
        r_cnt+=1

    wb.save(path_xls)
    wb.close()

    end_time = time.time()
    loguru.logger.info('>> Take time : '+f'{round(end_time-start_time,2)}'+'(S) => '+ \
                            f'{round((end_time-start_time)/60,2)}'+'(min).')


if __name__ == '__main__':

    with alive_bar(40, title='>> Generating...', length=40, bar='bubbles', spinner='radioactive') as bar:
        for i in range(40):
            time.sleep(.025)
            bar()

    loguru.logger.add(
        f'stock_datalog_daily_data_{datetime.date.today():%Y%m%d}.log',
        rotation='1 day',
        retention='7 days',
        level='DEBUG'
    )

    main()

    loguru.logger.info('       *******             *****        *             *     ***********                         *      ')
    loguru.logger.info('     *         *         *       *      * *           *    *                                   *       ')
    loguru.logger.info('     *          *       *         *     *  *          *    *                                  *        ')
    loguru.logger.info('     *           *     *           *    *   *         *    *                                 *         ')
    loguru.logger.info('     *            *    *           *    *    *        *    *                                *          ')
    loguru.logger.info('     *            *    *           *    *     *       *    *                               *           ')
    loguru.logger.info('     *            *    *           *    *      *      *     ***********      *            *            ')
    loguru.logger.info('     *            *    *           *    *       *     *    *                  *          *             ')
    loguru.logger.info('     *            *    *           *    *        *    *    *                   *        *              ')
    loguru.logger.info('     *           *     *           *    *         *   *    *                    *      *               ')
    loguru.logger.info('     *          *       *         *     *          *  *    *                     *    *                ')
    loguru.logger.info('     *         *         *       *      *           * *    *                      *  *                 ')
    loguru.logger.info('       *******             *****        *             *     ***********            *                   ')

    beep()
