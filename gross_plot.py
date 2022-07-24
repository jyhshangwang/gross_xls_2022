from turtle import color
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
import openpyxl
import gross_sub as sub
import loguru
import datetime


def stock_wscope_plot(title, lst, num, nam, **kwargs):

    plt.figure(dpi=300)
    plt.rcParams['font.sans-serif'] = ['Microsoft YaHei']

    plt.title(title, fontproperties=font1)
    plt.xlabel('日期(天)', fontproperties=font2)
    plt.ylabel('股價(元)', fontproperties=font2)
    Y_axis = lst
    X_axis = [ i+1 for i in range(len(Y_axis)) ]
    plt.plot(X_axis, Y_axis, label=str(num)+' '+nam)
    plt.legend(loc='best')
    #plt.show()
    plt.savefig('plot\\'+str(num)+'_'+nam+'.png')
    plt.close()
    loguru.logger.success(f'success: Save {str(num)}_{nam}.png file.')


def main():

    wb = sub.xls_wb_on(path_xls)
    st_lst = sub.xls_st_on(wb, 'Price')
    st_vol = sub.xls_st_on(wb, 'Volume')
    row_size = st_lst[0].max_row
    column_size = st_lst[1]

    para = input('Enter mode ( 0:single / 1:compare / 2:Price&Volume ) : ')


    while(int(para) == 0):
        #Stock_Num = input('Enter stock number : ')
        #Stock_Num = int(Stock_Num)
        #if Stock_Num == 0: break

        with open('gross_all_plot.txt', 'r') as fin:
            stocks = fin.readlines()

        for stock in stocks:
            for r in range(1,row_size+1,1):
                try:
                    if (st_lst[0].cell(row=r, column=2)).value == int(stock):
                        price_lst = []
                        for c in range(1,column_size+1,1):
                            price_lst.append((st_lst[0].cell(row=r, column=c)).value)
                        Stock_Name = price_lst[2]
                        price_lst = price_lst[4:]
                except Exception as e:
                    loguru.logger.error(e)
                #finally:
            stock_wscope_plot(date_str,price_lst,int(stock),Stock_Name)
        break


    if(int(para) == 1):
        stock_num_lst = []
        cnt=0
        while(1):
            Stock_Num = input('Enter the compared stock number : ')
            if int(Stock_Num) == 0:
                print(f'There are {str(cnt)} stock(s) adding to compare.')
                break
            stock_num_lst.append(int(Stock_Num))
            cnt+=1

        for n in range(len(stock_num_lst)):
            for r in range(1,row_size+1,1):
                if (st_lst[0].cell(row=r, column=2)).value == stock_num_lst[n]:
                    price_lst = []
                    for c in range(1,column_size+1,1):
                        price_lst.append((st_lst[0].cell(row=r, column=c)).value)
                    Y_axis = price_lst[4:]
                    X_axis = [ i+1 for i in range(len(Y_axis)) ]

                    plt.title('Compare', fontproperties=font1)
                    plt.xlabel('日期(天)', fontproperties=font2)
                    plt.ylabel('股價(元)', fontproperties=font2)
                    plt.plot(X_axis, Y_axis, label=str(stock_num_lst[n])+' '+price_lst[2])
                    plt.legend(loc='best')
        plt.savefig('compare.png')
        plt.close()


    if(int(para) == 2):
        with open('gross_all_plot.txt', 'r') as fin:
            stocks = fin.readlines()

        for stock in stocks:
            for r in range(1,row_size+1,1):
                if (st_lst[0].cell(row=r, column=2)).value == int(stock):
                    price_lst = []
                    volume_lst = []
                    for c in range(1,column_size+1,1):
                        price_lst.append((st_lst[0].cell(row=r, column=c)).value)
                        volume_lst.append((st_vol[0].cell(row=r, column=c)).value)
                    Stock_Name = price_lst[2]
                    Intro_info = price_lst[3]
                    Intro_info = Intro_info[0:10]+'...'
                    price_lst = price_lst[4:]
                    volume_lst = volume_lst[4:]
                    X_axis = [ i+1 for i in range(len(price_lst)) ]
                    fig, ax = plt.subplots(2,1, dpi=200)
                    plt.suptitle(str(int(stock))+'_'+str(Stock_Name)+'   '+Intro_info, fontproperties=font1)
                    #ax[0].set_title('股價(元)', fontproperties=font1)
                    ax[0].set_ylabel('股價(元)', fontproperties=font2)
                    ax[0].plot(X_axis, price_lst, color='navy')
                    #ax[1].set_title('成交量(張)', fontproperties=font1)
                    ax[1].set_ylabel('成交量(張)', fontproperties=font2)
                    ax[1].bar(X_axis, volume_lst, color='tomato')
                    #plt.show()
                    plt.tight_layout()
                    plt.savefig('plot\\'+str(int(stock))+'_'+str(Stock_Name)+'.png')
                    plt.close()
                    loguru.logger.success(f'success: Save {str(int(stock))}_{str(Stock_Name)}.png file.')

    sub.xls_wb_off(wb,path_xls)
    loguru.logger.success('success: End.')


if __name__ == '__main__':

    loguru.logger.add(
        f'stock_datalog_matplotlib_{datetime.date.today():%Y%m%d}.log',
        rotation='1 day',
        retention='7 days',
        level='DEBUG'    
    )

    fontPath = r'C:\\Users\\JS Wang\\AppData\\Roaming\\Python\\Python310\\site-packages\\matplotlib\\mpl-data\\fonts\\ttf\\NotoSansTC-Black.otf'
    font1 = fm.FontProperties(fname=fontPath, size=16)
    font2 = fm.FontProperties(fname=fontPath, size=10)
    path_xls = 'C:\\Users\\JS Wang\\Desktop\\test\\test.xlsx'

    date_str = sub.get_stock_datetime()

    #plt.figure(dpi=300)
    plt.rcParams['font.sans-serif'] = ['Microsoft YaHei']

    main()