import os
import time
import datetime
import random
import loguru
import openpyxl
from openpyxl.styles import Font
from bs4 import BeautifulSoup
import grequests
import requests
import chardet
import pyquery
import pandas as pd
import gross_class as cls


def time_title():
    print('\n===================================')
    print(' Time : '+str(datetime.datetime.today()))
    print('===================================')


def rand_on():
    delay_lst = [0.1, 0.2, 0.3, 0.4 ,0.5]
    delay = random.choice(delay_lst)
    time.sleep(delay)


def get_stock_datetime():
    reqs_date = requests.get("https://dj.mybank.com.tw/z/zc/zcl/zcl_2330.djhtm")

    if reqs_date.status_code != 200:
        loguru.logger.error('REQS: status code is not 200')
    loguru.logger.success('REQS: success')

    soup_date = BeautifulSoup(reqs_date.text, 'html.parser')
    date_tmp = (((soup_date.find('table', {'class': 't01'})).find_all('tr')[7]).find_all('td')[0]).text
    return date_tmp


def xls_wb_on(path_xls):
    return openpyxl.load_workbook(path_xls) if os.path.exists(path_xls) else openpyxl.Workbook()


def xls_wb_off(obj,path_xls):
    obj.save(path_xls)
    obj.close()


def xls_st_on(obj,st_name):
    flg=0
    for stn in obj.sheetnames: flg+=1 if stn == st_name else +0
    sheet = obj[st_name] if flg == 1 else obj.create_sheet(st_name,-1)
    column_cnt = sheet.max_column
    return [sheet, column_cnt]


def get_stock_urls(Stock_Num):
    urls = []
    #urls.append(f'http://kgieworld.moneydj.com/ZXQW/zc/zca/zca_{Stock_Num}.djhtm')
    #urls.append(f'https://kgieworld.moneydj.com/z/zc/zca/zca_{Stock_Num}.djhtm')
    #urls.append(f'http://jsjustweb.jihsun.com.tw/z/zc/zca/zca_{Stock_Num}.djhtm')
    urls.append(f'https://dj.mybank.com.tw/z/zc/zca/zca_{Stock_Num}.djhtm')
    return urls


@loguru.logger.catch
def get_reqs_data(urls):
    return [ requests.get(link) for link in urls ]


@loguru.logger.catch
def get_reqs_data_asynch(urls):
    reqs = ( grequests.get(link) for link in urls )
    response = grequests.imap(reqs, grequests.Pool(len(urls)))
    return response


@loguru.logger.catch
def parse_stock_data(reqs):
    for r in reqs:
        soup = BeautifulSoup(r.text,'html.parser')
        blocks = soup.find_all('table', {'class': 't01'})
        for blk in blocks:
            dat_p = float((((blk.find_all('tr')[1]).find_all('td')[7].text)).replace(',',''))
            dat_v = float((((blk.find_all('tr')[3]).find_all('td')[7].text)).replace(',',''))
            dat_c = float((((blk.find_all('tr')[13]).find_all('td')[1].text)).replace(',',''))
            dat_t = round((dat_v/dat_c/100),2)
    return dat_p,dat_v,dat_t


@loguru.logger.catch
def parse_stock_data_asynch(response,length):
    idnum_lst = []
    price_lst = []
    volum_lst = []
    tnrat_lst = []
    progress = cls.ProgressBar(length)
    for r in response:
        soup = BeautifulSoup(r.text,'html.parser')
        num_id = int(((soup.find('title')).text).replace('個股基本資料-',''))
        blocks = soup.find_all('table', {'class': 't01'})
        for blk in blocks:
            dat_p = float((((blk.find_all('tr')[1]).find_all('td')[7].text)).replace(',',''))
            dat_v = float((((blk.find_all('tr')[3]).find_all('td')[7].text)).replace(',',''))
            dat_c = float((((blk.find_all('tr')[13]).find_all('td')[1].text)).replace(',',''))
            dat_t = round((dat_v/dat_c/100),2)
        idnum_lst.append(num_id)
        price_lst.append(dat_p)
        volum_lst.append(dat_v)
        tnrat_lst.append(dat_t)
        progress.update()
    return idnum_lst,price_lst,volum_lst,tnrat_lst


@loguru.logger.catch
def parse_stock_data_yahoo(reqs): # for checking the price
    for r in reqs:
        soup = BeautifulSoup(r.text,'html.parser')
        if   soup.find('span', {'class': 'Fz(32px) Fw(b) Lh(1) Mend(16px) D(f) Ai(c) C($c-trend-up)'}) is not None :
            Yahoo_Price = (soup.find('span', {'class': 'Fz(32px) Fw(b) Lh(1) Mend(16px) D(f) Ai(c) C($c-trend-up)'})).text
        elif soup.find('span', {'class': 'Fz(32px) Fw(b) Lh(1) Mend(16px) D(f) Ai(c) C($c-trend-down)'}) is not None:
            Yahoo_Price = (soup.find('span', {'class': 'Fz(32px) Fw(b) Lh(1) Mend(16px) D(f) Ai(c) C($c-trend-down)'})).text
        elif soup.find('span', {'class': 'Fz(32px) Fw(b) Lh(1) Mend(16px) D(f) Ai(c)'}) is not None:
            Yahoo_Price = (soup.find('span', {'class': 'Fz(32px) Fw(b) Lh(1) Mend(16px) D(f) Ai(c)'})).text
        elif soup.find('span', {'class': 'Fz(32px) Fw(b) Lh(1) Mend(16px)'}) is not None:
            Yahoo_Price = (soup.find('span', {'class': 'Fz(32px) Fw(b) Lh(1) Mend(16px)'})).text
        elif soup.find('span', {'class': 'Fz(32px) Fw(b) Lh(1) Mend(16px) C(#fff) Px(6px) Py(2px) Bdrs(4px) Bgc($c-trend-up)'}) is not None:
            Yahoo_Price = (soup.find('span', {'class': 'Fz(32px) Fw(b) Lh(1) Mend(16px) C(#fff) Px(6px) Py(2px) Bdrs(4px) Bgc($c-trend-up)'})).text
        elif soup.find('span', {'class': 'Fz(32px) Fw(b) Lh(1) Mend(16px) C(#fff) Px(6px) Py(2px) Bdrs(4px) Bgc($c-trend-down)'}) is not None:
            Yahoo_Price = (soup.find('span', {'class': 'Fz(32px) Fw(b) Lh(1) Mend(16px) C(#fff) Px(6px) Py(2px) Bdrs(4px) Bgc($c-trend-down)'})).text
        #print(str(Yahoo_Price))
    return Yahoo_Price


@loguru.logger.catch
def parse_yahoo_asynch(response,length):
    idnum_lst = []
    price_lst = []
    pgs = cls.ProgressBar(length)
    for r in response:
        soup = BeautifulSoup(r.text,'html.parser')
        if   soup.find('span', {'class': 'Fz(32px) Fw(b) Lh(1) Mend(16px) D(f) Ai(c) C($c-trend-up)'}) is not None :
            price_lst.append((soup.find('span', {'class': 'Fz(32px) Fw(b) Lh(1) Mend(16px) D(f) Ai(c) C($c-trend-up)'})).text)
        elif soup.find('span', {'class': 'Fz(32px) Fw(b) Lh(1) Mend(16px) D(f) Ai(c) C($c-trend-down)'}) is not None:
            price_lst.append((soup.find('span', {'class': 'Fz(32px) Fw(b) Lh(1) Mend(16px) D(f) Ai(c) C($c-trend-down)'})).text)
        elif soup.find('span', {'class': 'Fz(32px) Fw(b) Lh(1) Mend(16px) D(f) Ai(c)'}) is not None:
            price_lst.append((soup.find('span', {'class': 'Fz(32px) Fw(b) Lh(1) Mend(16px) D(f) Ai(c)'})).text)
        elif soup.find('span', {'class': 'Fz(32px) Fw(b) Lh(1) Mend(16px)'}) is not None:
            price_lst.append((soup.find('span', {'class': 'Fz(32px) Fw(b) Lh(1) Mend(16px)'})).text)
        elif soup.find('span', {'class': 'Fz(32px) Fw(b) Lh(1) Mend(16px) C(#fff) Px(6px) Py(2px) Bdrs(4px) Bgc($c-trend-up)'}) is not None:
            price_lst.append((soup.find('span', {'class': 'Fz(32px) Fw(b) Lh(1) Mend(16px) C(#fff) Px(6px) Py(2px) Bdrs(4px) Bgc($c-trend-up)'})).text)
        elif soup.find('span', {'class': 'Fz(32px) Fw(b) Lh(1) Mend(16px) C(#fff) Px(6px) Py(2px) Bdrs(4px) Bgc($c-trend-down)'}) is not None:
            price_lst.append((soup.find('span', {'class': 'Fz(32px) Fw(b) Lh(1) Mend(16px) C(#fff) Px(6px) Py(2px) Bdrs(4px) Bgc($c-trend-down)'})).text)

        if soup.find('span', {'class': 'C($c-icon) Fz(24px) Mend(20px)'}) is not None:
            idnum_lst.append(int((soup.find('span', {'class': 'C($c-icon) Fz(24px) Mend(20px)'})).text))
        pgs.update()
    return idnum_lst , price_lst


@loguru.logger.catch
def cal_avg_price(obj,lst,row):
    sum_lst = [ 0 for i in range(len(lst)) ]
    avg_lst = [ 0 for i in range(len(lst)) ]
    for cnt in range(len(lst)):
        for clm in range(obj.max_column,obj.max_column-lst[cnt],-1): sum_lst[cnt]+=(obj.cell(row=row, column=clm)).value
        avg_lst[cnt]=round(sum_lst[cnt]/lst[cnt],2)
    return avg_lst


@loguru.logger.catch
def cal_moving_average_tangled(lst):
    cnt=0
    for i in range(len(lst)):
        if i == len(lst)-1: break
        for j in range(i+1,len(lst)):
            if lst[i]/lst[j] >= 0.95 and lst[i]/lst[j] <= 1.05: cnt+=1
    cmt = 'Yes' if cnt == 15 else '-'
    return cmt


@loguru.logger.catch
def cal_increase_rate(obj,lst,row):
    pri_lst = [ 0 for i in range(len(lst)) ]
    rat_lst = [ 0 for i in range(len(lst)) ]
    today_price = (obj.cell(row=row, column=obj.max_column)).value
    for cnt in range(len(lst)):
        pri_lst[cnt]=(obj.cell(row=row, column=obj.max_column-lst[cnt])).value
        rat_lst[cnt]=round(float((today_price-pri_lst[cnt])/pri_lst[cnt]*100),2) if pri_lst[cnt] != 0 else 0
    return rat_lst


@loguru.logger.catch
def cal_slope_rate(obj,r):
    d1 = float((obj.cell(row=r, column=obj.max_column-0)).value)
    d2 = float((obj.cell(row=r, column=obj.max_column-5)).value)
    slp_rat = round(((d1-d2)/d2)*100/5,2)
    return slp_rat


@loguru.logger.catch
def cal_price_position(obj1,obj2,r,nam):
    tday_price = float((obj1.cell(row=r, column=obj1.max_column)).value)
    line_price = float((obj2.cell(row=r, column=obj2.max_column)).value)
    if   tday_price  > line_price : pos_cmt = 'Yes'
    elif tday_price == line_price : pos_cmt = 'equal'
    elif tday_price  < line_price : pos_cmt = '-'
    return pos_cmt


@loguru.logger.catch
def cal_value_increase_rate(obj1,r):
    col = obj1.max_column
    sum_of_03days=0
    sum_of_20days=0
    val_lst = [ float((obj1.cell(row=r, column=i)).value) for i in range(col,col-23,-1) ]
    for i in range(0, 3): sum_of_03days+=val_lst[i]
    for i in range(3,23): sum_of_20days+=val_lst[i]
    vrate = round(((sum_of_03days/3)/(sum_of_20days/20)),2) if sum_of_20days != 0 else 0
    return vrate


@loguru.logger.catch
def check_reqs_data(path):
    reqs = requests.get(path)
    if reqs.status_code != 200:
        loguru.logger.error('REQS: status code is not 200.')
        return
    loguru.logger.success('REQS: success.')

    txt = None
    det = chardet.detect(reqs.content) # dict
    try:
        if det['confidence'] > 0.5:
            if det['encoding'] == 'big-5':
                txt = reqs.content.decode('big5')
            else:
                txt = reqs.content.decode(det['encoding'])
        else:
            txt = reqs.content.decode('utf-8')
    except Exception as e:
        loguru.logger.error(e) #try代碼塊出錯則會創建Exception類(class)對象，對象名為e，e中封裝了出錯的錯誤訊息
    if txt is None: return
    #loguru.logger.info(txt)
    return txt


@loguru.logger.catch
def dayily_info(path):

    txt = check_reqs_data(path)

    proportions = []

    d = pyquery.PyQuery(txt)
    tbs = list(d('table').items())
    tbs = tbs[2:3]
    for tb in tbs:
        trs = list(tb('tr').items())
        trs = trs[1:]
        for tr in trs:
            if tr == trs[0]:
                tds0 = list(tr('td').items())
                op_price = float((tds0[1].text().strip()).replace(',',''))
                hi_price = float((tds0[3].text().strip()).replace(',',''))
                lo_price = float((tds0[5].text().strip()).replace(',',''))
                td_price = float((tds0[7].text().strip()).replace(',',''))
            if tr == trs[1]:
                tds1 = list(tr('td').items())
                up_down = float((tds1[1].text().strip()).replace(',',''))
                hi_price_1y = float((tds1[3].text().strip()).replace(',',''))
                lo_price_1y = float((tds1[5].text().strip()).replace(',',''))
            if tr == trs[2]:
                tds2 = list(tr('td').items())
                pe_ratio = float((tds2[1].text().strip()).replace(',',''))
                mx_volume_1y = float((tds2[3].text().strip()).replace(',',''))
                mi_volume_1y = float((tds2[5].text().strip()).replace(',',''))
                td_volume = float((tds2[7].text().strip()).replace(',',''))
            if tr == trs[6]:
                tds6 = list(tr('td').items())
                incr_year = (tds6[1].text().strip().replace(',',''))
            if tr == trs[12]:
                tds12 = list(tr('td').items())
                stk_count = float((tds12[1].text().strip()).replace(',',''))
            if tr == trs[19]:
                tds19 = list(tr('td').items())
                Rev_rat_cmt = str((tds19[1].text().strip()).replace('、',' '))

        #proportions.append(cls.ProportionDailyInfo(td_price,td_volume))
        proportions.append(cls.ProportionDailyInfo(op_price,hi_price,lo_price,td_price,up_down,hi_price_1y,lo_price_1y,pe_ratio,mx_volume_1y,mi_volume_1y,td_volume,incr_year,stk_count,Rev_rat_cmt))

    message = os.linesep.join([str(prop) for prop in proportions])
    loguru.logger.info('Today:' + os.linesep + message)


@loguru.logger.catch
def revenue_info(path):

    txt = check_reqs_data(path)

    rev_propotions = []
    STK_REV = []
    rev_m = []
    rev_s = []
    rev_r = [1,1,10/8,1,1,1,1]
    rev_y = []

    d = pyquery.PyQuery(txt)
    trs = list(d('table tr').items())
    trs = trs[9:16]
    for tr in trs:
        tds = list(tr('td').items())
        code = tds[1].text().strip()
        if code != '':
            month = tds[0].text().strip()
            reven = tds[1].text().strip()
            mom   = tds[2].text().strip()
            yoy   = tds[4].text().strip()
            tyoy  = tds[6].text().strip()
            rev_propotions.append(cls.ProportionRevenueInfo(month,reven,mom,yoy,tyoy))
            rev_m.append(cls.ProportionRevenueInfo(month,reven,mom,yoy,tyoy).get_revenue())
            rev_y.append(cls.ProportionRevenueInfo(month,reven,mom,yoy,tyoy).get_yoyrate())

    #message = os.linesep.join([str(prop) for prop in rev_propotions])
    #loguru.logger.info('Today:' + os.linesep + message)

    for i in range(len(rev_m)): rev_s.append(rev_m[i]*rev_r[i])
    if( (rev_m[3]+rev_m[4]+rev_m[5]) == 0 or (rev_m[4]+rev_m[5]+rev_m[6]) == 0 ):
        STK_REV.append('NA')
        STK_REV.append('NA')
    else:
        DVO1 = (round((((rev_s[0]+rev_s[1]+rev_s[2])-(rev_s[3]+rev_s[4]+rev_s[5]))/(rev_s[3]+rev_s[4]+rev_s[5])),2))*100
        DVO2 = (round((((rev_s[1]+rev_s[2]+rev_s[3])-(rev_s[4]+rev_s[5]+rev_s[6]))/(rev_s[4]+rev_s[5]+rev_s[6])),2))*100
        DVRMS = round((DVO1-DVO2),2)
        STK_REV.append(DVO1)
        STK_REV.append(DVRMS)

    if  ( rev_s[0] > rev_s[1] and rev_s[1] > rev_s[2] and rev_s[2] > rev_s[3] ): REVCMT = 'INC 3'
    elif( rev_s[0] > rev_s[1] and rev_s[1] > rev_s[2] and rev_s[2] < rev_s[3] ): REVCMT = 'INC 2'
    elif( rev_s[0] > rev_s[1] and rev_s[1] < rev_s[2] and rev_s[2] < rev_s[3] ): REVCMT = 'INC 1'
    elif( rev_s[0] < rev_s[1] and rev_s[1] > rev_s[2] and rev_s[2] > rev_s[3] ): REVCMT = 'DEC -1'
    elif( rev_s[0] < rev_s[1] and rev_s[1] < rev_s[2] and rev_s[2] > rev_s[3] ): REVCMT = 'DEC -2'
    elif( rev_s[0] < rev_s[1] and rev_s[1] < rev_s[2] and rev_s[2] < rev_s[3] ): REVCMT = 'DEC -3'
    else:                                                                        REVCMT = 'NA'
    STK_REV.append(REVCMT)
    tmp = str(rev_propotions[0]).split(';')
    for i in range(len(tmp)-1): STK_REV.append(tmp[i])

    cnt=0
    for i in range(len(rev_y)-4):
        if( rev_y[i] < rev_y[i+1] ): cnt+=1
    if  ( cnt == 3             ): STK_REV.append('down')
    elif( cnt == 2 or cnt == 1 ): STK_REV.append('-')
    elif( cnt == 0             ): STK_REV.append('up')
