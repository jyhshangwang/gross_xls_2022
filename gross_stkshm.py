import openpyxl
import gross_sub as sub


def display(path_xls):

    wb = openpyxl.load_workbook(path_xls)
    st0 = sub.xls_st_on(wb,0,'Price'   ,0)
    st1 = sub.xls_st_on(wb,0,'Volume'  ,1)
    st2 = sub.xls_st_on(wb,0,'Value'   ,2)
    st3 = sub.xls_st_on(wb,0,'Turnover',3)

    CNT_JS=1
    CNT_LP=1
    print('\n 代號    價格   3日線   5日線   2周線    月線    成交值   3日漲幅   周漲幅  2周漲幅   月漲幅      股票                ')
    print('--------------------------------------------------------------------------------------------------------------------------------')
    for js0 in range(2,st2.max_row+1):

        val_lst=[0,0,0,0]
        CNT_VAL=0
        for js1 in range(len(val_lst)):
            val_lst[js1]=(st2.cell(row=js0, column=(st2.max_column-js1))).value
            if( val_lst[js1] > 1 ): CNT_VAL+=1
        if ( CNT_VAL >= 3 ): # >> 1.

            trn_lst=[0,0,0,0]
            CNT_TRN=0
            for js2 in range(1):
                trn_lst[js2]=(st3.cell(row=js0, column=st3.max_column)).value
                if( trn_lst[js2] > 1 ): CNT_TRN+=1
            if( CNT_TRN >= 1 ): # >> 2.

                day_lst=[ 3, 5,10,20]
                avg_lst=[ 0, 0, 0, 0]
                rat_lst=[ 0, 0, 0, 0]
                td_price=(st0.cell(row=js0, column=st0.max_column)).value
                avg_lst=sub.cal_avg_price(st0,day_lst,js0)
                rat_lst=sub.cal_incr_rate(st0,day_lst,js0)
                cal_tmp=0
                for js3 in range(len(day_lst)): cal_tmp+=1 if ((td_price/avg_lst[js3] > 0.9) and (td_price/avg_lst[js3] < 1.8)) else +0 #3. diff ratio

                if( rat_lst[3] > -10 ): #4. Month increase rate
                    if cal_tmp == 4:
                        #if ( (avg_lst[0] > avg_lst[1]) and (avg_lst[1] > avg_lst[2]) and (avg_lst[2] > avg_lst[3]) and (td_price < 500) ):
                        if ( (avg_lst[0] >= avg_lst[1]) and (td_price < 300) ):
                            if( CNT_LP%15 == 0 ): print(' 代號    價格   3日線   5日線   2周線    月線    成交值   3日漲幅   周漲幅  2周漲幅   月漲幅      股票                ')
                            if( CNT_LP%15 == 0 ): print('--------------------------------------------------------------------------------------------------------------------------------')
                            print('%5s'%str((st0.cell(row=js0, column=2)).value), end='')
                            print('%8s'%str(td_price), end='')
                            print('%8s'%str(round(avg_lst[0],2)), end='')
                            print('%8s'%str(round(avg_lst[1],2)), end='')
                            print('%8s'%str(round(avg_lst[2],2)), end='')
                            print('%8s'%str(round(avg_lst[3],2)), end='')
                            print('%8s'%str((st2.cell(row=js0, column=st2.max_column)).value)+str('億'), end='')
                            print('%9s'%str(rat_lst[0])+'% ', end='')
                            print('%7s'%str(rat_lst[1])+'% ', end='')
                            print('%7s'%str(rat_lst[2])+'% ', end='')
                            print('%7s'%str(rat_lst[3])+'% ', end='')
                            print('{:>5}'.format(str((st0.cell(row=js0, column=3)).value)[:4]), end='')
                            print('{:>9}'.format(str((st0.cell(row=js0, column=4)).value)[:8]), end='')
                            print(' >>> 近3日正在啟動!', end='') if( rat_lst[0]/rat_lst[1] > 0.8 ) else print(end='')
                            print()
                            CNT_LP+=1
        CNT_JS+=1
    wb.save(path_xls)

