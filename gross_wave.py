import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
import openpyxl
import gross_sub as sub
import loguru
import datetime


fontPath = r'C:\\Users\\JS Wang\\AppData\\Roaming\\Python\\Python310\\site-packages\\matplotlib\\mpl-data\\fonts\\ttf\\NotoSansTC-Black.otf'
font1 = fm.FontProperties(fname=fontPath, size=16)
font2 = fm.FontProperties(fname=fontPath, size=10)
path_xls = 'C:\\Users\\JS Wang\\Desktop\\test\\test.xlsx'

date_str = sub.get_stock_datetime()

plt.figure(dpi=300)
plt.rcParams['font.sans-serif'] = ['Microsoft YaHei']

WSCOPE1 = 0
WSCOPE2 = 1


def stock_wscope_plot(title, lst, num, nam, **kwargs):

    plt.figure(dpi=300)
    plt.rcParams['font.sans-serif'] = ['Microsoft YaHei']

    plt.title(title, fontproperties=font1)
    plt.xlabel('日期(天)', fontproperties=font2)
    plt.ylabel('股價(元)', fontproperties=font2)
    Y_axis = lst
    X_axis = [ i+1 for i in range(len(Y_axis)) ]
    plt.plot(X_axis, Y_axis, label=str(num)+' '+nam)
    plt.legend(loc='lower left')
    plt.savefig(str(num)+'_'+nam+'.png')
    plt.close()
    loguru.logger.success('success: Store the stock waveform.')



# >> Compare waveform ..........................................................
if WSCOPE1 == 1:
    wb = sub.xls_wb_on(path_xls)
    st_lst = sub.xls_st_on(wb,'Price')

    for r in range(st_lst[0].max_row-1):
        data_price = []
        date_time = []
        for c in range(st_lst[1]):
            if r == 0: date_time.append((st_lst[0].cell(row=1, column=c+1)).value)
            data_price.append((st_lst[0].cell(row=r+2, column=c+1)).value)
        date_time = date_time[4:]
        Y_axis = data_price[4:]
        X_axis = [ i+1 for i in range(len(Y_axis)) ]

        plt.title('Compare 2022/07/15', fontproperties=font1)
        plt.xlabel('日期(天)', fontproperties=font2)
        plt.ylabel('股價(元)', fontproperties=font2)
        plt.plot(X_axis, Y_axis, label=str(data_price[1])+' '+data_price[2])
        plt.legend(loc = 'lower left')
        #plt.show()

    plt.savefig('1522.png')
    plt.close()

    sub.xls_wb_off(wb,path_xls)


# >> stock waveform ..........................................................
if WSCOPE2 == 1:

    while(True):
        Stock_Num = input('Enter stock number : ')
        Stock_Num = int(Stock_Num)
        if Stock_Num == 0: break

        wb = sub.xls_wb_on(path_xls)
        st_lst = sub.xls_st_on(wb,'Price')
        row_size = st_lst[0].max_row
        column_size = st_lst[1]
        for r in range(1,row_size+1,1):
            try:
                if (st_lst[0].cell(row=r, column=2)).value == Stock_Num:
                    price_lst = []
                    for c in range(1,column_size+1,1):
                        price_lst.append((st_lst[0].cell(row=r, column=c)).value)
                    Stock_Name = price_lst[2]
                    price_lst = price_lst[4:]
            except Exception as e:
                loguru.logger.error(e)
            #finally:
        stock_wscope_plot(date_str,price_lst,Stock_Num,Stock_Name)
        sub.xls_wb_off(wb,path_xls)

