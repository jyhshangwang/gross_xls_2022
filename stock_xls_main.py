import loguru
import time
import datetime
import pandas as pd
import gross_sub as sub
import gross_stkshm as shm
import gross_yahoo as yas
import gross_class as cls
from openpyxl.styles import Font
import winsound
from alive_progress import alive_bar


# ======        ======
# ======  main  ======
# ======        ======
if __name__ == '__main__':

    with alive_bar(50, title='>> Starting stock program', length=40, bar='blocks') as bar:
        for i in range(50):
            time.sleep(.025)
            bar()

    loguru.logger.add(
        f'stock_datalog_{datetime.date.today():%Y%m%d}.log',
        rotation='1 day',
        retention='7 days',
        level='DEBUG'
    )
    para = input('\n>> Enter parameters : ')
    HIDAR_EXCEL = [int(para[0]),int(para[1])]
    YSTK_M = int(para[2])
    FK_LST = [0,0,1]

    sub.time_title()
    start_time = time.time()
    date_tmp = sub.get_stock_datetime()

    path_xls = 'C:\\Users\\JS Wang\\Desktop\\test\\tmp.xlsx'
    path_xls_out = 'C:\\Users\\JS Wang\\Desktop\\test\\output.xlsx'
    path_fin = 'C:\\Users\\JS Wang\\Desktop\\test\\gross_all_0115.txt'

    if YSTK_M: yas.yahoo_stock_data()
    if FK_LST[0]: sub.counter_info('http://jsjustweb.jihsun.com.tw/z/zc/zcl/zcl_3006.djhtm')#sub.revenue_info('https://dj.mybank.com.tw/z/zc/zch/zch_3006.djhtm')
    if FK_LST[1]: shm.display(path_xls)


    if HIDAR_EXCEL[0]:

        fi_stk = open(path_fin,'r')
        lines = fi_stk.readlines()

        STK_CNT = 1
        STK_PRI = []
        STK_VOL = []
        STK_TRR = []
        STK_PRI.insert(0,date_tmp)
        STK_VOL.insert(0,date_tmp)
        STK_TRR.insert(0,date_tmp)

        if( FK_LST[2] == 0 ):
            JS_TMP = [0,0,0]
            for line in lines:
                STK_NUM = int(line)
                JS_TMP = sub.parse_stock_data(sub.get_reqs_data(sub.get_stock_urls(str(STK_NUM))))
                STK_PRI.append(float(JS_TMP[0]))
                STK_VOL.append(float(JS_TMP[1]))
                STK_TRR.append(float(JS_TMP[2]))
                print(">> No."+str(STK_CNT) + "  ...  "+str(STK_NUM))
                #sub.rand_on()
                STK_CNT+=1

        if( FK_LST[2] == 1 ):
            idlst = []
            links = []
            JS_TMP2 = [0,0,0,0]
            idlst = [ int(line) for line in lines ]
            links = [ f'https://dj.mybank.com.tw/z/zc/zca/zca_{str(int(line))}.djhtm' for line in lines ]
            JS_TMP2 = sub.parse_stock_data_asynch(sub.get_reqs_data_asynch(links),len(idlst))
            loguru.logger.info('Sort the data now ...')
            for cnt_i in range(0,len(idlst)):
                tmpid = idlst[cnt_i]
                for cnt_j in range(0,len(JS_TMP2[0])):
                    if JS_TMP2[0][cnt_j] == tmpid:
                        print('\r'+str(JS_TMP2[0][cnt_j]), end='')
                        STK_PRI.append(float(JS_TMP2[1][cnt_j]))
                        STK_VOL.append(float(JS_TMP2[2][cnt_j]))
                        STK_TRR.append(float(JS_TMP2[3][cnt_j]))
                        break
            print()

        loguru.logger.info('>> STEP1. Write to  ... '+str(path_xls))
        step1_dict = {
            '0':'Price',
            '1':'Volume',
            '2':'Value',
            '3':'Turnover'
        }
        wb = sub.xls_wb_on(path_xls)
        step1_lst = [ sub.xls_st_on(wb,step1_dict[str(i)]) for i in range(len(step1_dict)) ]

        for r1 in range(1,len(STK_PRI)+1): (step1_lst[0][0].cell(row=r1, column=step1_lst[0][1]+1)).value = STK_PRI[r1-1]
        for r2 in range(1,len(STK_VOL)+1): (step1_lst[1][0].cell(row=r2, column=step1_lst[1][1]+1)).value = STK_VOL[r2-1]
        for r3 in range(1,len(STK_PRI)+1): (step1_lst[2][0].cell(row=r3, column=step1_lst[2][1]+1)).value = STK_PRI[r3-1] if r3 == 1 else round((float(STK_PRI[r3-1])*float(STK_VOL[r3-1])/100000),2)
        for r4 in range(1,len(STK_TRR)+1): (step1_lst[3][0].cell(row=r4, column=step1_lst[3][1]+1)).value = STK_TRR[r4-1]

        loguru.logger.success('Completion OK: Capture daily info.')
        sub.xls_wb_off(wb,path_xls)


    if HIDAR_EXCEL[1]:

        # >> STEP2 ..............................................................................................
        loguru.logger.info('>> STEP2. Calculate average line price from  ... '+str(path_xls))
        step2_dict = {
            '0':'Price',
            '1':'3ma',
            '2':'5ma',
            '3':'10ma',
            '4':'20ma',
            '5':'40ma',
            '6':'60ma',
            '7':'Tangled'
        }
        wb = sub.xls_wb_on(path_xls)
        step2_lst = [ sub.xls_st_on(wb,step2_dict[str(i)]) for i in range(len(step2_dict)) ]

        pgs2 = cls.ProgressBar(step2_lst[0][0].max_row)
        for r in range(1,step2_lst[0][0].max_row+1):
            day_lst=[ 3, 5,10,20,40,60]
            avg_lst=[ 0, 0, 0, 0, 0, 0]
            if r != 1 :
                avg_lst=sub.cal_avg_price(step2_lst[0][0],day_lst,r)
                avg_cmt=sub.cal_moving_average_tangled(avg_lst)
                #avg_lst.append(avg_cmt)
            pgs2.update()
            for j in range(len(avg_lst)): (step2_lst[j+1][0].cell(row=r, column=step2_lst[j+1][1]+1)).value = avg_lst[j] if r != 1 else date_tmp
            (step2_lst[7][0].cell(row=r, column=step2_lst[7][1]+1)).value = avg_cmt if r != 1 else date_tmp
        loguru.logger.success('Completion OK: Average line price')
        sub.xls_wb_off(wb,path_xls)


        # >> STEP3 ..............................................................................................
        loguru.logger.info('>> STEP3. Calculate Increase rate from  ... '+str(path_xls))
        step3_dict = {
            '0':'Price',
            '1':'Inc1',
            '2':'Inc3',
            '3':'Inc5',
            '4':'Inc10',
            '5':'Inc20',
            '6':'Inc60'
        }
        wb = sub.xls_wb_on(path_xls)
        step3_lst = [ sub.xls_st_on(wb,step3_dict[str(i)]) for i in range(len(step3_dict)) ]

        pgs3 = cls.ProgressBar(step3_lst[0][0].max_row)
        for r in range(1,step3_lst[0][0].max_row+1):
            day_lst=[ 1, 3, 5,10,20,60]
            rat_lst=[ 0, 0, 0, 0, 0, 0]
            if r!= 1 : rat_lst=sub.cal_increase_rate(step3_lst[0][0],day_lst,r)
            pgs3.update()
            for j in range(len(rat_lst)): (step3_lst[j+1][0].cell(row=r, column=step3_lst[j+1][1]+1)).value = rat_lst[j] if r != 1 else date_tmp
        loguru.logger.success('Completion OK: Increase rate')
        sub.xls_wb_off(wb,path_xls)


        # >> STEP4 ..............................................................................................
        loguru.logger.info('>> STEP4. Calculate slope from  ... '+str(path_xls))
        step4_dict = {
            '0':'20ma',
            '1':'Slope20',
            '2':'60ma',
            '3':'Slope60'
        }
        wb = sub.xls_wb_on(path_xls)
        step4_lst = [ sub.xls_st_on(wb,step4_dict[str(i)]) for i in range(len(step4_dict)) ]

        pgs4 = cls.ProgressBar(step4_lst[0][0].max_row)
        for r in range(1,step4_lst[0][0].max_row+1):
            if r != 1:
                val1=sub.cal_slope_rate(step4_lst[0][0],r)
                val2=sub.cal_slope_rate(step4_lst[2][0],r)
            pgs4.update()
            (step4_lst[1][0].cell(row=r, column=step4_lst[1][1]+1)).value = val1 if r != 1 else date_tmp
            (step4_lst[3][0].cell(row=r, column=step4_lst[3][1]+1)).value = val2 if r != 1 else date_tmp
        loguru.logger.success('Completion OK: Slope value')
        sub.xls_wb_off(wb,path_xls)


        # >> STEP5 ..............................................................................................
        loguru.logger.info('>> STEP5. Calculate price vs avg line price ... '+str(path_xls))
        step5_dict = {
            '0':'Price',
            '1':'20ma',
            '2':'20CMT',
            '3':'60ma',
            '4':'60CMT'
        }
        wb = sub.xls_wb_on(path_xls)
        step5_lst = [ sub.xls_st_on(wb,step5_dict[str(i)]) for i in range(len(step5_dict)) ]

        pgs5 = cls.ProgressBar(step5_lst[0][0].max_row)
        for r in range(1,step5_lst[0][0].max_row+1):
            if r != 1 :
                cmt_tmp1 = str(sub.cal_price_position(step5_lst[0][0],step5_lst[1][0],r,'20ma'))
                cmt_tmp2 = str(sub.cal_price_position(step5_lst[0][0],step5_lst[3][0],r,'60ma'))
            pgs5.update()
            (step5_lst[2][0].cell(row=r, column=step5_lst[2][1]+1)).value = str(cmt_tmp1) if r != 1 else date_tmp
            (step5_lst[4][0].cell(row=r, column=step5_lst[4][1]+1)).value = str(cmt_tmp2) if r != 1 else date_tmp
        loguru.logger.success('Completion OK: Price vs Avg line price relation')
        sub.xls_wb_off(wb,path_xls)


        # >> STEP6 ..............................................................................................
        loguru.logger.info('>> STEP6. Calculate the change in value in 3 days ... '+str(path_xls))
        step6_dict = {
            '0':'Value',
            '1':'Vrate'
        }
        wb = sub.xls_wb_on(path_xls)
        step6_lst = [ sub.xls_st_on(wb,step6_dict[str(i)]) for i in range(len(step6_dict)) ]

        pgs6 = cls.ProgressBar(step6_lst[0][0].max_row)
        for r in range(1,step6_lst[0][0].max_row+1):
            if r != 1 : vrt_tmp = sub.cal_value_increase_rate(step6_lst[0][0],r)
            pgs6.update()
            (step6_lst[1][0].cell(row=r, column=step6_lst[1][1]+1)).value = vrt_tmp if r != 1 else date_tmp
        loguru.logger.success('Completion OK: Value rate')
        sub.xls_wb_off(wb,path_xls)


        # >> STEP7 ..............................................................................................
        loguru.logger.info('>> STEP7. Combine today data in the same sheet ... ')
        wb = sub.xls_wb_on(path_xls)
        wb_out = sub.xls_wb_on(path_xls_out)
        wb_out.remove(wb_out['Today'])
        st_out = sub.xls_st_on(wb_out,'Today')

        tmp_clm = st_out[0].max_column
        st_dict = {
                    'Price':'股價', 'Volume':'成交量', 'Value':'成交值', 'Turnover':'周轉率',
                    '3ma':'3日線', '5ma':'5日線', '10ma':'10日線', '20ma':'20日線', '40ma':'40日線', '60ma':'60日線',
                    'Inc1':'1日漲幅', 'Inc3':'3日漲幅', 'Inc5':'5日漲幅', 'Inc10':'10日漲幅', 'Inc20':'20日漲幅', 'Inc60':'60日漲幅',
                    'Slope20':'月線斜率', 'Slope60':'季線斜率', '20CMT':'站上月線?', '60CMT':'站上季線?', 'Tangled':'均線糾結?', 'Vrate':'值增率', 'Force':'主力'
                }
        pgs7 = cls.ProgressBar(len(st_dict))
        for nam in wb.sheetnames:
            st = wb[nam]
            lst_tmp = []
            for i in range(st.max_row): lst_tmp.append(0)
            for r in range(1,st.max_row+1): lst_tmp[r-1] = ((st.cell(row=r , column=st.max_column)).value) if r != 1 else st_dict[nam]
            for r in range(1,st.max_row+1):
                (st_out[0].cell(row=r, column=tmp_clm+1)).font = Font(name='Calibri', size=12)
                (st_out[0].cell(row=r, column=tmp_clm+1)).value = lst_tmp[r-1]
            tmp_clm+=1
            pgs7.update()
        loguru.logger.success('Completion OK: Combination')
        sub.xls_wb_off(wb,path_xls)
        sub.xls_wb_off(wb_out,path_xls_out)


    end_time = time.time()
    loguru.logger.info('\n運算時間 : '+str(round((end_time-start_time),2))+'(S) >> '+str(round(round((end_time-start_time),2)/60,2))+'(min)')


    duration = 1000 # mS
    freq = 400      # Hz
    for i in range(5):
        if i%2 == 0 : winsound.Beep(freq, duration)
        time.sleep(0.2)

    print()
    print("       *******             *****        *             *     ***********                         *      ")
    print("     *         *         *       *      * *           *    *                                   *       ")
    print("     *          *       *         *     *  *          *    *                                  *        ")
    print("     *           *     *           *    *   *         *    *                                 *         ")
    print("     *            *    *           *    *    *        *    *                                *          ")
    print("     *            *    *           *    *     *       *    *                               *           ")
    print("     *            *    *           *    *      *      *     ***********      *            *            ")
    print("     *            *    *           *    *       *     *    *                  *          *             ")
    print("     *            *    *           *    *        *    *    *                   *        *              ")
    print("     *           *     *           *    *         *   *    *                    *      *               ")
    print("     *          *       *         *     *          *  *    *                     *    *                ")
    print("     *         *         *       *      *           * *    *                      *  *                 ")
    print("       *******             *****        *             *     ***********            *                   ")
    print()
